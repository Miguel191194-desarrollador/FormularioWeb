# app.py
from flask import Flask, render_template, request, redirect, flash
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import io
import threading
import base64
import tempfile
import os
import logging
import requests

# (Opcional en local) Cargar .env si existe; en Render usará Environment
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

app = Flask(__name__)
app.secret_key = 'supersecretkey'

# Logging claro a consola (Render captura stdout)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

# =========================
#  CONFIG (ENTORNO/RENDER)
# =========================
GAS_WEBHOOK_URL = os.getenv("GAS_WEBHOOK_URL")   # URL de Apps Script (termina en /exec)
MAIL_TO_ADMIN   = os.getenv("MAIL_TO_ADMIN")     # opcional: copia
# Umbral para dividir correo si adjuntos son pesados (por defecto 18 MiB)
EMAIL_SPLIT_THRESHOLD_BYTES = int(os.getenv("EMAIL_SPLIT_THRESHOLD_BYTES", str(18 * 1024 * 1024)))
# Forzar envío síncrono para depurar (muestra error en la respuesta del navegador)
FORCE_SYNC_SEND = os.getenv("FORCE_SYNC_SEND", "false").lower() in ("1", "true", "yes")

# =========================
#  RUTAS
# =========================

@app.route('/', methods=['GET'])
def formulario():
    return render_template('formulario.html')

@app.route('/plantas', methods=['POST', 'GET'])
def plantas():
    if request.method == 'GET':
        flash('Por favor, rellena primero el formulario de cliente.')
        return redirect('/')
    datos_cliente = request.form.to_dict()
    return render_template('plantas.html', datos_cliente=datos_cliente)

@app.route('/guardar', methods=['POST'])
def guardar():
    form_data = request.form.to_dict()
    plantas_data = request.form.to_dict()
    data = {**form_data, **plantas_data}

    # Firma en base64 (proveniente de un <canvas> o similar)
    firma_base64 = data.get('firma_cliente')
    firma_bytes = None
    if firma_base64:
        try:
            # Se espera "data:image/png;base64,AAAA..."
            firma_bytes = base64.b64decode(firma_base64.split(",")[1])
        except Exception:
            firma_bytes = None

    # Validación: al menos una planta
    hay_una_planta = any(plantas_data.get(f'planta_nombre_{i}') for i in range(1, 11))
    if not hay_una_planta:
        flash('⚠️ Debes rellenar al menos los datos de una planta antes de continuar.')
        return render_template('plantas.html', datos_cliente=form_data)

    # Generar excels en memoria
    try:
        archivo_excel_cliente = crear_excel_en_memoria(data, firma_bytes)
        archivo_excel_plantas = crear_excel_plantas_en_memoria(data)
    except Exception as e:
        logging.exception("❌ Error generando Excels")
        flash(f'Error generando Excels: {e}')
        return render_template('plantas.html', datos_cliente=form_data)

    # Comprobación temprana del webhook
    if not GAS_WEBHOOK_URL:
        logging.error("❌ GAS_WEBHOOK_URL no está configurado en el entorno")
        flash('Error de configuración: falta GAS_WEBHOOK_URL en el servidor.')
        return render_template('gracias.html')

    nombre_cliente = data.get('nombre') or "cliente"
    correo_comercial = data.get('correo_comercial')

    # Envío en primer plano (depuración) o en hilo (producción)
    if FORCE_SYNC_SEND:
        ok, detalle = enviar_excel_con_posible_division(
            archivo_excel_cliente, archivo_excel_plantas,
            correo_comercial, nombre_cliente
        )
        if ok:
            flash('Documentación enviada correctamente.')
        else:
            flash(f'Error enviando la documentación: {detalle}')
        return render_template("gracias.html")
    else:
        threading.Thread(
            target=_thread_enviar,
            args=(archivo_excel_cliente, archivo_excel_plantas, correo_comercial, nombre_cliente),
            daemon=True
        ).start()
        return render_template("gracias.html")

def _thread_enviar(archivo1, archivo2, correo_comercial, nombre_cliente):
    try:
        ok, detalle = enviar_excel_con_posible_division(
            archivo1, archivo2, correo_comercial, nombre_cliente
        )
        if ok:
            logging.info("✅ Envío completado: %s", detalle)
        else:
            logging.error("❌ Fallo de envío: %s", detalle)
    except Exception as e:
        logging.exception("❌ Excepción enviando correo en hilo: %s", e)

# =========================
#  FUNCIONES EXCEL
# =========================

def crear_excel_en_memoria(data, firma_bytes=None):
    """
    Rellena la plantilla 'Copia de Alta de Cliente.xlsx' (hoja 'FICHA CLIENTE')
    y devuelve un BytesIO con el archivo.
    """
    wb = load_workbook("Copia de Alta de Cliente.xlsx")
    ws = wb["FICHA CLIENTE"]

    ws["B4"] = data.get("nombre")
    ws["B5"] = data.get("nif")
    ws["D5"] = data.get("telefono_general")
    ws["B6"] = data.get("email_general")
    ws["D6"] = data.get("web")
    ws["B7"] = data.get("direccion")
    ws["D7"] = data.get("cp")
    ws["B8"] = data.get("poblacion")
    ws["D8"] = data.get("provincia")
    ws["B13"] = data.get("forma_pago")
    ws["B18"] = data.get("compras_nombre")
    ws["D18"] = data.get("compras_telefono")
    ws["B19"] = data.get("compras_email")
    ws["B22"] = data.get("contabilidad_nombre")
    ws["D22"] = data.get("contabilidad_telefono")
    ws["B24"] = data.get("contabilidad_email")
    ws["B27"] = data.get("facturacion_nombre")
    ws["D27"] = data.get("facturacion_telefono")
    ws["B29"] = data.get("facturacion_email")
    ws["B32"] = data.get("descarga_nombre")
    ws["D32"] = data.get("descarga_telefono")
    ws["B34"] = data.get("descarga_email")
    ws["C38"] = data.get("contacto_documentacion")
    ws["C39"] = data.get("contacto_devoluciones")
    ws["B43"] = data.get("sepa_nombre_banco")
    ws["B44"] = data.get("sepa_domicilio_banco")
    ws["B45"] = data.get("sepa_cp")
    ws["B46"] = data.get("sepa_poblacion")
    ws["B47"] = data.get("sepa_provincia")
    ws["B48"] = data.get("iban_completo")

    # Insertar imagen de firma si la hay
    if firma_bytes:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            tmp.write(firma_bytes)
            tmp_path = tmp.name
        try:
            img = ExcelImage(tmp_path)
            img.width = 200
            img.height = 60
            ws.add_image(img, "B49")
        finally:
            try:
                os.remove(tmp_path)
            except Exception:
                pass

    excel_mem = io.BytesIO()
    wb.save(excel_mem)
    excel_mem.seek(0)
    return excel_mem

def crear_excel_plantas_en_memoria(data):
    """
    Rellena la plantilla 'Copia de Alta de Plantas.xlsx' (hoja 'Plantas')
    y devuelve un BytesIO con el archivo.
    Empieza a escribir en fila 4 (3 + i) y usa columnas B..M (orden directo).
    """
    wb = load_workbook("Copia de Alta de Plantas.xlsx")
    ws = wb["Plantas"]

    columnas = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    campos = [
        "planta_nombre_{}", "planta_direccion_{}", "planta_cp_{}", "planta_poblacion_{}",
        "planta_provincia_{}", "planta_telefono_{}", "planta_email_{}", "planta_horario_{}",
        "planta_observaciones_{}", "planta_contacto_nombre_{}", "planta_contacto_telefono_{}",
        "planta_contacto_email_{}"
    ]

    for i in range(1, 11):
        fila = 3 + i  # i=1 -> fila 4
        valores = [data.get(campo.format(i), "") for campo in campos]
        if not valores[0]:  # sin nombre de planta -> omite
            continue
        for col, valor in zip(columnas, valores):
            ws[f"{col}{fila}"] = valor

    excel_mem = io.BytesIO()
    wb.save(excel_mem)
    excel_mem.seek(0)
    return excel_mem

# =========================
#  ENVÍO POR WEBHOOK GMAIL
# =========================

def _build_recipients(correo_comercial):
    """Construye lista de destinatarios."""
    destinatarios = ['tesoreria@dimensasl.com']
    if correo_comercial and "@" in correo_comercial:
        destinatarios.append(correo_comercial)
    if MAIL_TO_ADMIN and "@" in MAIL_TO_ADMIN:
        destinatarios.append(MAIL_TO_ADMIN)
    return destinatarios

def _encode_attachment(bytes_io, filename):
    """Devuelve dict de adjunto para Apps Script a partir de BytesIO."""
    bytes_io.seek(0)
    raw = bytes_io.getvalue()
    return {
        "filename": filename,
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "base64": base64.b64encode(raw).decode("utf-8"),
        "raw_size": len(raw)  # útil para logs/decisión de split
    }

def _post_to_webhook(payload):
    """Hace POST al webhook y devuelve (ok, detalle)."""
    try:
        r = requests.post(GAS_WEBHOOK_URL, json=payload, timeout=30)
        logging.info("Webhook status=%s body=%s", r.status_code, r.text[:500])
        if r.status_code == 200 and "OK" in r.text:
            return True, f"status={r.status_code}"
        return False, f"Webhook error status={r.status_code} body={r.text[:500]}"
    except Exception as e:
        logging.exception("Excepción en requests.post")
        return False, f"Excepción: {e}"

def enviar_excel_con_posible_division(archivo1, archivo2, correo_comercial, nombre_cliente):
    """
    Prepara adjuntos y envía:
      - Un único correo con ambos adjuntos si caben bajo el umbral.
      - O dos correos (uno por adjunto) si superan el umbral.
    """
    if not GAS_WEBHOOK_URL:
        return False, "Falta GAS_WEBHOOK_URL"

    destinatarios = _build_recipients(correo_comercial)
    to_csv = ",".join(destinatarios)

    subj_base = f"Alta de cliente: {nombre_cliente} — Documentación"
    body_html = construir_body_html(nombre_cliente)

    att1 = _encode_attachment(archivo1, f"Alta Cliente - {nombre_cliente}.xlsx")
    att2 = _encode_attachment(archivo2, f"Alta Plantas - {nombre_cliente}.xlsx")

    # Estimación del tamaño en base64 (≈ 4/3)
    total_raw = att1["raw_size"] + att2["raw_size"]
    total_b64_est = int(total_raw * 4 / 3)

    logging.info("Tamaño adjuntos (raw): %d bytes; estimado base64: %d; umbral: %d",
                 total_raw, total_b64_est, EMAIL_SPLIT_THRESHOLD_BYTES)

    if total_b64_est <= EMAIL_SPLIT_THRESHOLD_BYTES:
        # Enviar un único correo con los dos adjuntos
        payload = {
            "to": to_csv,
            "subject": subj_base,
            "text": "Alta de cliente — Documentación adjunta",
            "html": body_html,
            "attachments": [
                {k: v for k, v in att1.items() if k != "raw_size"},
                {k: v for k, v in att2.items() if k != "raw_size"},
            ]
        }
        ok, detalle = _post_to_webhook(payload)
        return ok, ("1 correo (2 adjuntos). " + detalle)
    else:
        # Enviar dos correos: uno por cada Excel
        payload1 = {
            "to": to_csv,
            "subject": f"{subj_base} (1/2) — Copia Alta de Cliente",
            "text": "Adjunto el Excel de la primera página (Cliente).",
            "html": body_html,
            "attachments": [{k: v for k, v in att1.items() if k != "raw_size"}]
        }
        ok1, det1 = _post_to_webhook(payload1)

        payload2 = {
            "to": to_csv,
            "subject": f"{subj_base} (2/2) — Copia Alta de Plantas",
            "text": "Adjunto el Excel de la segunda página (Plantas).",
            "html": body_html,
            "attachments": [{k: v for k, v in att2.items() if k != "raw_size"}]
        }
        ok2, det2 = _post_to_webhook(payload2)

        ok_total = ok1 and ok2
        detalle = f"Split en 2 correos. 1/2: {det1} | 2/2: {det2}"
        return ok_total, detalle

def construir_body_html(nombre_cliente):
    """HTML del cuerpo del correo (completo, con tablas de riesgos/sector/subsector)."""
    return f"""
    <html>
    <body>
    <p>Buenas,</p>
    <p>Se ha completado el alta de un nuevo cliente en el sistema: <strong>{nombre_cliente}</strong>.</p>
    <p>Adjuntamos en este correo dos archivos Excel:<br>
    - Uno con los datos generales del cliente.<br>
    - Otro con la información detallada de sus plantas.</p>

    <p><strong><span style='color:red;'>⚠️ IMPORTANTE: REENVIAR ESTE CORREO A MIGUEL INDICANDO EL RIESGO A SOLICITAR PARA ESTE CLIENTE, SECTOR Y SUBSECTOR.</span></strong></p>

    <p><strong>Seleccione el riesgo, el sector y el subsector marcando la casilla correspondiente:</strong></p>

    <table style="width: 100%; border-collapse: collapse;" cellspacing="15">
        <tr>
            <td style="vertical-align: top;">
                <table style="border-collapse: collapse; border: 1px solid black;">
                    <thead>
                        <tr><th style="padding: 5px; border: 1px solid black;">Riesgo</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                    </thead>
                    <tbody>
                        <tr><td style="padding:5px; border:1px solid black;">0</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">500</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">1000</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">1500</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">2000</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">2500</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">3000</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">3500</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">4000</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">4500</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">5000</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">20000</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">Otro (especificar)</td><td style="padding:5px; border:1px solid black;"><input type="text" placeholder="Escriba aquí el riesgo"></td></tr>
                    </tbody>
                </table>
            </td>

            <td style="vertical-align: top;">
                <table style="border-collapse: collapse; border: 1px solid black;">
                    <thead>
                        <tr><th style="padding: 5px; border: 1px solid black;">Sector</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                    </thead>
                    <tbody>
                        <tr><td style="padding:5px; border:1px solid black;">Agricultura</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">Aguas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">Alimentación</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">Distribuidor</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">Ganadería</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">Industrial</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">Piscinas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">Sector0</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                    </tbody>
                </table>
            </td>

            <td style="vertical-align: top;">
                <table style="border-collapse: collapse; border: 1px solid black;">
                    <thead>
                        <tr><th colspan="2" style="padding: 5px; border: 1px solid black;">Subsectores</th></tr>
                    </thead>
                    <tbody>
                        <tr><th style="padding: 5px; border: 1px solid black;">Agricultura</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AG)Agricultura</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><th style="padding: 5px; border: 1px solid black;">Aguas</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(A)Industrial</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(A)Potable</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(A)Residual</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><th style="padding: 5px; border: 1px solid black;">Alimentación</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Aceituna</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Aditivos, aromas, azucares y salsas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Bebidas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Cárnicas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Chocolate, café y confiteria</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Conserva - procesado frutas, hortalizas y cereales</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Grasas animales y vegetales</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Lácteos</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Panadería,pasta,harina,galletas, y pasteleria</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Pescado</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(AL)Vino</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><th style="padding: 5px; border: 1px solid black;">Distribuidor</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(D)Agricultura</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(D)Aguas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(D)Alimentación</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(D)Ganadería</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(D)Industrial</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(D)Piscinas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><th style="padding: 5px; border: 1px solid black;">Ganadería</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(G)Explotaciones Ganaderas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(G)Fabricación Alimentos FEED</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><th style="padding: 5px; border: 1px solid black;">Industrial</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Biodiésel</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Cemento,yeso y hormigón</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Comercio</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Construcción</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Detergencia y Cosmética</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Energía</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Energía Renovable</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Farmacia</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Fertilizantes y agroquímicos</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Madera</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Metalurgia</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Minerales</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Papel y cartón</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Petróleo y gas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Pinturas,barnices,resinas,masillas,tintas</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Plástico</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Química básica</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Química fina / formulados</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Residuos</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Textil y curtidos</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Transportes</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(I)Vidrio y Cerámica</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><th style="padding: 5px; border: 1px solid black;">Piscinas</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(P)Privada</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(P)Pública</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                        <tr><th style="padding: 5px; border: 1px solid black;">Sector 0</th><th style="padding: 5px; border: 1px solid black;">Selección</th></tr>
                        <tr><td style="padding:5px; border:1px solid black;">(S)Sector 0</td><td style="padding:5px; border:1px solid black;"><input type="checkbox"></td></tr>
                    </tbody>
                </table>
            </td>
        </tr>
    </table>

    <p>Gracias por vuestra colaboración.</p>
    <p>Un saludo,<br>Departamento de Tesorería</p>
    </body>
    </html>
    """

# =========================
#  MAIN
# =========================

if __name__ == '__main__':
    # En local: python app.py
    # En Render: el propio servicio lanza el proceso con tu start command.
    app.run(debug=True)


